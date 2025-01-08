import pandas as pd
import json
import os
import re
import requests
from datetime import datetime
from urllib.parse import urlparse, parse_qs
from openpyxl.utils import get_column_letter
import numpy as np

class ExcelToJsonConverter:
    def __init__(self, excel_file, output_folder="output", images_folder="images"):
        self.excel_file = excel_file
        self.output_folder = output_folder
        self.images_folder = os.path.join(output_folder, images_folder)
        self.columns = [
            "Бот", "Нынешний юз", "Первый запуск", "Второй запуск", 
            "Старый юз", "Кд", "Примечания", "Дата смерти", "Владелец",
            "Функционал (команды)", "Кол-во юзеров", "Статус",
            "Каналы Ботов", "Сурсы", "Айди"
        ]
        self._setup_folders()

    def _setup_folders(self):
        os.makedirs(self.output_folder, exist_ok=True)
        os.makedirs(self.images_folder, exist_ok=True)

    def _normalize_spaces(self, text):
        if not isinstance(text, str):
            return text
        return ' '.join(text.split())

    def _extract_drive_id(self, url):
        if not url or not isinstance(url, str):
            return None
        
        if 'drive.google.com' not in url:
            return None

        try:
            if 'file/d/' in url:
                file_id = url.split('file/d/')[1].split('/')[0]
                return file_id
            else:
                parsed = urlparse(url)
                query_params = parse_qs(parsed.query)
                file_id = query_params.get('id', [None])[0]
                return file_id
        except Exception as e:
            print(f"Error extracting drive ID from URL {url}: {e}")
            return None

    def _download_image(self, file_id, row_index):
        if not file_id:
            return None

        download_url = f"https://drive.usercontent.google.com/u/0/uc?id={file_id}&export=download"
        
        try:
            response = requests.get(download_url)
            if response.status_code == 200:
                file_path = os.path.join(self.images_folder, f"image_{row_index}_{file_id}.jpg")
                with open(file_path, 'wb') as f:
                    f.write(response.content)
                return file_path
        except Exception as e:
            print(f"Error downloading image: {e}")
        
        return None

    def _convert_to_text(self, value):
        if pd.isna(value):
            return "Пустая ячейка"
        
        if isinstance(value, str):
            return self._normalize_spaces(value.strip())
        
        if isinstance(value, (pd.Timestamp, datetime)):
            return value.strftime("%d %B %Y").replace("January", "января")\
                                            .replace("February", "февраля")\
                                            .replace("March", "марта")\
                                            .replace("April", "апреля")\
                                            .replace("May", "мая")\
                                            .replace("June", "июня")\
                                            .replace("July", "июля")\
                                            .replace("August", "августа")\
                                            .replace("September", "сентября")\
                                            .replace("October", "октября")\
                                            .replace("November", "ноября")\
                                            .replace("December", "декабря")
        
        if isinstance(value, (int, float)):
            if value.is_integer():
                return str(int(value))
            return str(value)
        
        return self._normalize_spaces(str(value).strip())

    def _get_cell_hyperlink(self, worksheet, row_idx, col_idx):
        try:
            col_letter = get_column_letter(col_idx + 1)
            cell = worksheet[f"{col_letter}{row_idx + 2}"]  # +2 because we skip header and Excel is 1-based
            if cell.hyperlink:
                return cell.hyperlink.target
        except Exception as e:
            print(f"Error getting hyperlink from cell {col_letter}{row_idx + 2}: {e}")
        return None

    def convert(self):
        # Read Excel file with modified parameters
        df = pd.read_excel(
            self.excel_file,
            dtype=str,
            na_filter=False,
            keep_default_na=False,
            header=0  # Explicitly set first row as header
        )
        
        # Print debug info
        print("Processing Excel file...")
        print(f"Found {len(df)} rows")
        print("Columns:", df.columns.tolist())
        
        # Open workbook to access hyperlinks
        from openpyxl import load_workbook
        wb = load_workbook(self.excel_file)
        ws = wb.active
        
        # Ensure all columns exist
        for col in self.columns:
            if col not in df.columns:
                print(f"Column {col} not found in Excel file")
                df[col] = "Пустая ячейка"

        # Process cells and create result
        result = []
        for idx, row in df.iterrows():
            entry = {}
            for col in self.columns:
                col_idx = df.columns.get_loc(col)
                
                # Get cell value and normalize spaces
                value = self._convert_to_text(row[col])
                
                # Get hyperlink if exists
                hyperlink = self._get_cell_hyperlink(ws, idx, col_idx)
                
                if hyperlink and 'drive.google.com' in hyperlink:
                    print(f"Found hyperlink in row {idx+1}, column {col}: {hyperlink}")
                    file_id = self._extract_drive_id(hyperlink)
                    if file_id:
                        image_path = self._download_image(file_id, idx)
                        entry[col] = {
                            'text': value,
                            'link': hyperlink,
                            'image_path': image_path
                        }
                    else:
                        entry[col] = value
                else:
                    entry[col] = value
            
            result.append(entry)

        # Close workbook
        wb.close()

        # Save JSON
        output_file = os.path.join(self.output_folder, 'output4.json')
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)

        return output_file

def main():
    converter = ExcelToJsonConverter('table.xlsx')
    output_file = converter.convert()
    print(f"Conversion completed. Output saved to: {output_file}")

if __name__ == "__main__":
    main()